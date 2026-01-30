import openpyxl

def create_sample_excel(filename="sample.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Data"
    
    # Add some headers
    ws['A1'] = "ID"
    ws['B1'] = "Name"
    ws['C1'] = "Value"
    
    # Add some data
    data = [
        (1, "Item A", 100),
        (2, "Item B", 200),
        (3, "Item C", 300),
    ]
    
    for row in data:
        ws.append(row)
        
    wb.save(filename)
    print(f"Created {filename}")

if __name__ == "__main__":
    create_sample_excel()
