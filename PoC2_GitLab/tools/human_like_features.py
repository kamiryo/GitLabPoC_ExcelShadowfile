"""
human_like_features.py

人間が作成したような特徴をExcelファイルに付与するモジュール
"""

import random
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# 業務でよく使われる色 (Excelのテーマカラーと標準色)
COMMON_COLORS = [
    'FFFFFF',  # White
    'FFFF00',  # Yellow (highlight)
    'C6E0B4',  # Light Green
    'BDD7EE',  # Light Blue
    'FFE699',  # Light Orange
    'F4B084',  # Light Coral
    'E2EFDA',  # Very Light Green
    'D9E1F2',  # Very Light Blue
]

# 日本の業務でよく使われるフォント
COMMON_FONTS = [
    'MS ゴシック',
    'MS 明朝', 
    'メイリオ',
    'Yu Gothic',
    'Arial',
    'Calibri',
]


def apply_random_column_widths(worksheet, min_width=8, max_width=25, variation=0.15):
    """
    列幅にランダムなばらつきを追加
    
    Args:
        worksheet: openpyxl worksheet object
        min_width: 最小列幅
        max_width: 最大列幅  
        variation: ばらつきの程度 (0.0-1.0)
    """
    for col_idx in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        
        # ベース幅を計算（セルの内容に基づく）
        max_length = 0
        for cell in worksheet[col_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # 基本幅を設定
        base_width = min(max(max_length + 2, min_width), max_width)
        
        # ランダムなばらつきを追加
        random_factor = 1 + random.uniform(-variation, variation)
        adjusted_width = base_width * random_factor
        adjusted_width = min(max(adjusted_width, min_width), max_width)
        
        worksheet.column_dimensions[col_letter].width = adjusted_width


def apply_random_row_heights(worksheet, base_height=15, variation=0.1):
    """
    行の高さに微妙なばらつきを追加
    
    Args:
        worksheet: openpyxl worksheet object
        base_height: ベースとなる行の高さ
        variation: ばらつきの程度
    """
    for row_idx in range(1, worksheet.max_row + 1):
        # 10% の確率で高さを変更
        if random.random() < 0.1:
            random_factor = 1 + random.uniform(-variation, variation)
            adjusted_height = base_height * random_factor
            worksheet.row_dimensions[row_idx].height = adjusted_height


def apply_natural_formatting(worksheet, header_row=1, probability=0.3):
    """
    自然な書式設定を適用（フォント、太字、色など）
    
    Args:
        worksheet: openpyxl worksheet object
        header_row: ヘッダー行の番号（この行には必ず書式を適用）
        probability: 各セルに書式を適用する確率
    """
    # ヘッダー行の書式設定（高確率で太字・背景色）
    for cell in worksheet[header_row]:
        if cell.value:
            # 太字
            cell.font = Font(
                name=random.choice(COMMON_FONTS[:3]),  # ゴシック系が多い
                size=random.choice([10, 11, 11, 12]),  # 11が多め
                bold=True
            )
            # 背景色（ヘッダーは色付き）
            if random.random() < 0.8:  # 80%の確率で色を付ける
                cell.fill = PatternFill(
                    start_color=random.choice(COMMON_COLORS[2:6]),
                    end_color=random.choice(COMMON_COLORS[2:6]),
                    fill_type='solid'
                )
            # 中央揃え
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # データ行の書式設定（ランダム）
    for row in worksheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            if cell.value and random.random() < probability:
                # フォント設定
                font_name = random.choice(COMMON_FONTS)
                font_size = random.choice([9, 10, 10, 11, 11, 11, 12])  # 10-11が多め
                is_bold = random.random() < 0.1  # 10%の確率で太字
                
                cell.font = Font(name=font_name, size=font_size, bold=is_bold)
                
                # 背景色（5%の確率）
                if random.random() < 0.05:
                    cell.fill = PatternFill(
                        start_color=random.choice(COMMON_COLORS),
                        end_color=random.choice(COMMON_COLORS),
                        fill_type='solid'
                    )


def add_typing_imperfections(worksheet, probability=0.15):
    """
    タイプミスや入力の不完全さを追加
    
    Args:
        worksheet: openpyxl worksheet object  
        probability: 各セルに不完全さを追加する確率
    """
    for row in worksheet.iter_rows(min_row=2):  # ヘッダー以外
        for cell in row:
            if cell.value and isinstance(cell.value, str) and random.random() < probability:
                value = str(cell.value)
                
                # ランダムに不完全さを追加
                imperfection_type = random.choice([
                    'leading_space',
                    'trailing_space', 
                    'double_space',
                    'zenkaku_space',
                    'mixed_width',
                ])
                
                if imperfection_type == 'leading_space':
                    # 前に空白を追加
                    cell.value = ' ' + value
                    
                elif imperfection_type == 'trailing_space':
                    # 後ろに空白を追加
                    cell.value = value + ' '
                    
                elif imperfection_type == 'double_space':
                    # ランダムな位置に2つのスペース
                    if len(value) > 2:
                        pos = random.randint(1, len(value) - 1)
                        cell.value = value[:pos] + '  ' + value[pos:]
                        
                elif imperfection_type == 'zenkaku_space':
                    # 全角スペースを混入
                    cell.value = value.replace(' ', '　', 1)
                    
                elif imperfection_type == 'mixed_width':
                    # 数字の全角半角混在
                    if any(c.isdigit() for c in value):
                        # 半角数字を全角に変換
                        zenkaku_map = str.maketrans('0123456789', '０１２３４５６７８９')
                        # ランダムに一部を全角化
                        chars = list(value)
                        for i, c in enumerate(chars):
                            if c.isdigit() and random.random() < 0.3:
                                chars[i] = c.translate(zenkaku_map)
                        cell.value = ''.join(chars)


def apply_natural_alignment(worksheet, header_row=1):
    """
    自然な配置を適用（左揃え、中央、右揃えの混在）
    
    Args:
        worksheet: openpyxl worksheet object
        header_row: ヘッダー行の番号
    """
    for row in worksheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            if cell.value:
                # 数値は右揃え、文字列は左揃え（基本）
                if isinstance(cell.value, (int, float)):
                    alignment = 'right'
                else:
                    # 文字列の場合、90%は左揃え、10%は中央
                    alignment = 'left' if random.random() < 0.9 else 'center'
                
                cell.alignment = Alignment(horizontal=alignment, vertical='center')


def apply_all_human_features(worksheet, header_row=1, intensity='medium'):
    """
    すべての人間らしい特徴を一括適用
    
    Args:
        worksheet: openpyxl worksheet object
        header_row: ヘッダー行の番号
        intensity: 適用の強度 ('light', 'medium', 'heavy')
    """
    # 強度に応じて確率を調整
    intensity_map = {
        'light': {'format': 0.1, 'imperfect': 0.05, 'col_var': 0.05, 'row_var': 0.05},
        'medium': {'format': 0.3, 'imperfect': 0.15, 'col_var': 0.15, 'row_var': 0.1},
        'heavy': {'format': 0.5, 'imperfect': 0.25, 'col_var': 0.25, 'row_var': 0.15},
    }
    
    params = intensity_map.get(intensity, intensity_map['medium'])
    
    # 各特徴を適用
    apply_random_column_widths(worksheet, variation=params['col_var'])
    apply_random_row_heights(worksheet, variation=params['row_var'])
    apply_natural_formatting(worksheet, header_row=header_row, probability=params['format'])
    add_typing_imperfections(worksheet, probability=params['imperfect'])
    apply_natural_alignment(worksheet, header_row=header_row)
    
    print(f"Applied human-like features with '{intensity}' intensity to worksheet '{worksheet.title}'")
