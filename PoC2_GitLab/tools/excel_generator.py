#!/usr/bin/env python3
"""
excel_generator.py

人間が作成したような特徴を持つExcelファイルを生成・加工するメインツール
"""

import sys
import argparse
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 同じディレクトリのモジュールをインポート
try:
    from human_like_features import apply_all_human_features
    from sample_data import (
        get_sample_data_generator,
        generate_employee_list,
        generate_sales_data,
        generate_project_schedule,
        generate_budget_table,
        generate_inventory_list,
    )
except ImportError:
    # スクリプトとして実行された場合のフォールバック
    import os
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from human_like_features import apply_all_human_features
    from sample_data import (
        get_sample_data_generator,
        generate_employee_list,
        generate_sales_data,
        generate_project_schedule,
        generate_budget_table,
        generate_inventory_list,
    )


class ExcelGenerator:
    """Excelファイル生成・加工クラス"""
    
    def __init__(self, filepath=None):
        """
        Args:
            filepath: 既存のExcelファイルパス（Noneの場合は新規作成）
        """
        self.filepath = filepath
        if filepath and Path(filepath).exists():
            self.workbook = load_workbook(filepath)
            print(f"Loaded existing workbook: {filepath}")
        else:
            self.workbook = Workbook()
            # デフォルトシートを削除
            if self.workbook.sheetnames:
                self.workbook.remove(self.workbook.active)
            print("Created new workbook")
    
    def create_sheet(self, sheet_name, data=None, header_row=1):
        """
        新しいシートを作成してデータを挿入
        
        Args:
            sheet_name: シート名
            data: 挿入するデータ（辞書のリスト）
            header_row: ヘッダー行の番号
            
        Returns:
            作成したワークシート
        """
        if sheet_name in self.workbook.sheetnames:
            print(f"Sheet '{sheet_name}' already exists. Removing and recreating.")
            self.workbook.remove(self.workbook[sheet_name])
        
        worksheet = self.workbook.create_sheet(sheet_name)
        
        if data:
            self._insert_data(worksheet, data, header_row)
        
        return worksheet
    
    def _insert_data(self, worksheet, data, header_row=1):
        """
        データをワークシートに挿入
        
        Args:
            worksheet: 対象のワークシート
            data: 挿入するデータ（辞書のリスト）
            header_row: ヘッダー行の番号
        """
        if not data:
            return
        
        # ヘッダーを挿入
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=header_row, column=col_idx, value=header)
        
        # データを挿入
        for row_idx, record in enumerate(data, start=header_row + 1):
            for col_idx, header in enumerate(headers, start=1):
                value = record.get(header)
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        print(f"Inserted {len(data)} rows into worksheet '{worksheet.title}'")
    
    def add_data_to_sheet(self, sheet_name, data, start_row=None):
        """
        既存のシートにデータを追加
        
        Args:
            sheet_name: シート名
            data: 追加するデータ（辞書のリスト）
            start_row: データを追加開始する行（Noneの場合は最終行の次）
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist")
        
        worksheet = self.workbook[sheet_name]
        
        if start_row is None:
            start_row = worksheet.max_row + 1
        
        # データ挿入（ヘッダーは既存のものを使用）
        if worksheet.max_row >= 1:
            headers = [cell.value for cell in worksheet[1]]
        else:
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=col_idx, value=header)
            start_row = 2
        
        for row_idx, record in enumerate(data, start=start_row):
            for col_idx, header in enumerate(headers, start=1):
                value = record.get(header)
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        print(f"Added {len(data)} rows to sheet '{sheet_name}' starting at row {start_row}")
    
    def apply_human_features(self, sheet_name=None, intensity='medium'):
        """
        人間らしい特徴を適用
        
        Args:
            sheet_name: 対象のシート名（Noneの場合は全シート）
            intensity: 適用の強度 ('light', 'medium', 'heavy')
        """
        if sheet_name:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' does not exist")
            apply_all_human_features(self.workbook[sheet_name], intensity=intensity)
        else:
            # 全シートに適用
            for ws in self.workbook.worksheets:
                apply_all_human_features(ws, intensity=intensity)
    
    def save(self, filepath=None):
        """
        ワークブックを保存
        
        Args:
            filepath: 保存先のパス（Noneの場合は既存のパスを使用）
        """
        if filepath is None:
            filepath = self.filepath
        
        if filepath is None:
            raise ValueError("No filepath specified for saving")
        
        # ディレクトリが存在しない場合は作成
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        
        self.workbook.save(filepath)
        print(f"Workbook saved to: {filepath}")


def create_command(args):
    """新規Excel作成コマンド"""
    generator = ExcelGenerator()
    
    # テンプレートに基づいてデータ生成
    if args.template:
        data_func = get_sample_data_generator(args.template)
        if data_func is None:
            print(f"Unknown template: {args.template}")
            print("Available templates: employee, sales, project, budget, inventory")
            sys.exit(1)
        
        # 行数指定があればそれを使用
        if args.rows:
            if args.template in ['employee', 'employee_list']:
                data = data_func(num_employees=args.rows)
            elif args.template in ['sales', 'sales_data']:
                data = data_func(num_records=args.rows)
            elif args.template in ['project', 'project_schedule', 'schedule']:
                data = data_func(num_tasks=args.rows)
            elif args.template in ['budget', 'budget_table']:
                data = data_func(num_items=args.rows)
            elif args.template in ['inventory', 'inventory_list']:
                data = data_func(num_items=args.rows)
            else:
                data = data_func()
        else:
            data = data_func()
        
        # シート名を決定
        sheet_name = args.sheet_name or args.template.capitalize()
        generator.create_sheet(sheet_name, data)
    else:
        # 空のシートを作成
        sheet_name = args.sheet_name or 'Sheet1'
        generator.create_sheet(sheet_name)
    
    # 人間らしい特徴を適用
    if not args.no_human_features:
        intensity = args.intensity or 'medium'
        generator.apply_human_features(intensity=intensity)
    
    # 保存
    output_path = args.output or 'output.xlsx'
    generator.save(output_path)
    
    print(f"\n✓ Excel file created successfully: {output_path}")


def modify_command(args):
    """既存Excel修正コマンド"""
    if not Path(args.file).exists():
        print(f"Error: File not found: {args.file}")
        sys.exit(1)
    
    generator = ExcelGenerator(args.file)
    
    # データ追加
    if args.add_rows:
        # デフォルトでは最初のシートに追加
        sheet_name = args.sheet_name or generator.workbook.sheetnames[0]
        
        # テンプレートからデータ生成
        if args.template:
            data_func = get_sample_data_generator(args.template)
            if data_func:
                data = data_func(num_employees=args.add_rows) if 'employee' in args.template \
                       else data_func(num_records=args.add_rows) if 'sales' in args.template \
                       else data_func(num_tasks=args.add_rows) if 'project' in args.template or 'schedule' in args.template \
                       else data_func(num_items=args.add_rows)
                
                generator.add_data_to_sheet(sheet_name, data)
    
    # 人間らしい特徴を再適用
    if args.randomize and not args.no_human_features:
        intensity = args.intensity or 'medium'
        sheet_name = args.sheet_name if args.sheet_name else None
        generator.apply_human_features(sheet_name=sheet_name, intensity=intensity)
    
    # 保存
    output_path = args.output or args.file
    generator.save(output_path)
    
    print(f"\n✓ Excel file modified successfully: {output_path}")


def test_command(args):
    """テストコマンド - 各テンプレートのサンプルを生成"""
    print("Running test mode - generating sample files for all templates...\n")
    
    templates = ['employee', 'sales', 'project', 'budget', 'inventory']
    
    for template in templates:
        print(f"Generating {template} sample...")
        generator = ExcelGenerator()
        
        data_func = get_sample_data_generator(template)
        data = data_func()
        
        generator.create_sheet(template.capitalize(), data)
        generator.apply_human_features(intensity='medium')
        
        output_path = f'test_{template}.xlsx'
        generator.save(output_path)
    
    print("\n✓ All test files generated successfully!")


def main():
    parser = argparse.ArgumentParser(
        description='人間らしいExcelファイル生成・加工ツール',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用例:
  # 新規作成
  python excel_generator.py create --template sales --output sales_data.xlsx
  python excel_generator.py create --template employee --rows 100 --output employees.xlsx
  
  # 既存ファイルの修正
  python excel_generator.py modify --file data.xlsx --add-rows 50 --template sales
  python excel_generator.py modify --file data.xlsx --randomize --intensity heavy
  
  # テスト
  python excel_generator.py test
        """
    )
    
    subparsers = parser.add_subparsers(dest='command', help='コマンド')
    
    # Create コマンド
    create_parser = subparsers.add_parser('create', help='新規Excelファイルを作成')
    create_parser.add_argument('--template', type=str, help='テンプレート名 (employee, sales, project, budget, inventory)')
    create_parser.add_argument('--rows', type=int, help='生成する行数')
    create_parser.add_argument('--output', '-o', type=str, help='出力ファイルパス')
    create_parser.add_argument('--sheet-name', type=str, help='シート名')
    create_parser.add_argument('--no-human-features', action='store_true', help='人間らしい特徴を適用しない')
    create_parser.add_argument('--intensity', choices=['light', 'medium', 'heavy'], help='人間らしさの強度')
    
    # Modify コマンド
    modify_parser = subparsers.add_parser('modify', help='既存Excelファイルを修正')
    modify_parser.add_argument('--file', '-f', type=str, required=True, help='対象ファイルパス')
    modify_parser.add_argument('--add-rows', type=int, help='追加する行数')
    modify_parser.add_argument('--template', type=str, help='追加データのテンプレート')
    modify_parser.add_argument('--sheet-name', type=str, help='対象シート名')
    modify_parser.add_argument('--randomize', action='store_true', help='人間らしい特徴を再適用')
    modify_parser.add_argument('--no-human-features', action='store_true', help='人間らしい特徴を適用しない')
    modify_parser.add_argument('--intensity', choices=['light', 'medium', 'heavy'], help='人間らしさの強度')
    modify_parser.add_argument('--output', '-o', type=str, help='出力ファイルパス（指定しない場合は上書き）')
    
    # Test コマンド
    test_parser = subparsers.add_parser('test', help='テストモード（全テンプレートのサンプル生成）')
    
    args = parser.parse_args()
    
    if args.command == 'create':
        create_command(args)
    elif args.command == 'modify':
        modify_command(args)
    elif args.command == 'test':
        test_command(args)
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
